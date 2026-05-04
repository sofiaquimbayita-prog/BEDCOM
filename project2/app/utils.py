"""
UTILIDADES PARA EXPORTACIÓN DE REPORTES
Módulo con funciones para exportar datos a PDF y Excel
"""

from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings
from django.contrib.staticfiles import finders
from django.http import HttpResponse
from django.template.loader import render_to_string
from datetime import datetime
import io
import base64
import re
from pathlib import Path
from urllib.parse import urlparse

# ====== UTILIDADES COMPARTIDAS ======
def sanitize_value(value):
    """Limpia texto para evitar saltos de línea y caracteres de escape en exportes."""
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.replace('\\r', ' ').replace('\\n', ' ').replace('\\t', ' ')
        text = text.replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
        text = re.sub(r'\\s{2,}', ' ', text).strip()
        text = re.sub(r'(.)\\1{2,}', r'\\1\\1', text)
        return text
    return value


def resolve_static_file_url(value):
    """Convierte URLs de /static/ a file:// para evitar peticiones HTTP lentas en WeasyPrint."""
    if not isinstance(value, str):
        return value

    parsed = urlparse(value)
    path = parsed.path or value
    static_url = settings.STATIC_URL

    if path.startswith(static_url):
        static_path = path[len(static_url):]
    elif path.startswith(f"/{static_url}"):
        static_path = path[len(static_url) + 1:]
    else:
        return value

    found_path = finders.find(static_path)
    if not found_path:
        return value

    return Path(found_path).resolve().as_uri()


def normalize_pdf_context(contexto):
    if contexto.get('logo_url'):
        contexto['logo_url'] = resolve_static_file_url(contexto['logo_url'])
    return contexto


# ====== EXPORTACIÓN A PDF ======
def exportar_pdf(titulo, columnas, datos, nombre_archivo, contexto_extra=None):
    """
    FUNCIÓN PARA EXPORTAR DATOS A PDF USANDO WEASYPRINT
    """
    # Crear contexto base para el template
    # Formatemos la fecha directamente como string para evitar problemas con filtros de Django
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    contexto = {
        'titulo': titulo,
        'columnas': columnas,
        'datos': [[sanitize_value(valor) if isinstance(valor, str) else valor for valor in fila] for fila in datos],
        'now': fecha_actual,
    }
    
    # Si se pasan datos adicionales (como la URL del icono), se agregan al contexto
    if contexto_extra:
        contexto.update(contexto_extra)

    contexto = normalize_pdf_context(contexto)
    
    # Generar HTML desde el template
    html_string = render_to_string('reportes/reporte_pdf.html', contexto)
    
    # Crear documento PDF usando base local para evitar resoluciones HTTP lentas
    html_object = HTML(string=html_string, base_url=Path(settings.BASE_DIR).resolve().as_uri())
    
    # Generar PDF en memoria
    pdf_bytes = html_object.write_pdf()
    
    # Crear respuesta HTTP con el PDF
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
    
    return response


# ====== EXPORTACIÓN A EXCEL ======
def exportar_excel(titulo, columnas, datos, nombre_archivo):
    """
    FUNCIÓN PARA EXPORTAR DATOS A EXCEL USANDO OPENPYXL
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Estilos para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    # Escribir título del reporte
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(columnas))
    title_cell = ws.cell(row=1, column=1, value=titulo)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = alignment

    # Escribir encabezados de columna
    for col_num, column_title in enumerate(columnas, 1):
        cell = ws.cell(row=3, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Escribir datos
    safe_datos = [
        [sanitize_value(valor) if isinstance(valor, str) else valor for valor in fila]
        for fila in datos
    ]

    for row_num, fila in enumerate(safe_datos, 4):
        for col_num, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_num, column=col_num, value=valor)
            cell.alignment = Alignment(horizontal="left")

    # Guardar en memoria y retornar respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    return response


# ====== EXPORTACIÓN DE ESTADÍSTICAS A PDF ======
def exportar_pdf_estadisticas(titulo, estadisticas, nombre_archivo, contexto_extra=None):
    """
    FUNCIÓN PARA EXPORTAR ESTADÍSTICAS GENERALES A PDF USANDO WEASYPRINT
    Muestra métricas clave del sistema en formato de tarjetas
    """
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    contexto = {
        'titulo': titulo,
        'estadisticas': estadisticas,
        'now': fecha_actual,
    }
    
    if contexto_extra:
        contexto.update(contexto_extra)

    contexto['estadisticas'] = [
        (
            sanitize_value(metric) if isinstance(metric, str) else metric,
            sanitize_value(value) if isinstance(value, str) else value,
        )
        for metric, value in estadisticas
    ]

    contexto = normalize_pdf_context(contexto)
    
    # Generar HTML desde el template especializado para estadísticas
    html_string = render_to_string('reportes/reporte_estadisticas.html', contexto)
    
    # Crear documento PDF
    html_object = HTML(string=html_string, base_url=Path(settings.BASE_DIR).resolve().as_uri())
    
    # Generar PDF en memoria
    pdf_bytes = html_object.write_pdf()
    
    # Crear respuesta HTTP con el PDF
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
    
    return response
